import telebot
import datetime
from datetime import datetime, timedelta
import pytz
import openpyxl
import threading
import time
import re
import os
import logging
import math

# Set up logging
logging.basicConfig(
    filename=r'F:\Python\bot_debug.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Bot configuration
TOKEN = '8393419861:AAGU7wRKcZFLCwkBmNVEl62edCcroTOKRgU'
CHAT_ID = 1150807213
bot = telebot.TeleBot(TOKEN)
FILE_PATH = r'F:\Python\Activity_log_py.xlsx'
TZ = pytz.timezone('Africa/Cairo')

# Global state variables
multi_end = None
current_slot_start = None
last_slot_start = None
has_response = False
timeout_timer = None

def get_current_slot_start(dt):
    """Get the start of the current 30-minute slot."""
    rounded = dt.replace(second=0, microsecond=0)
    minutes = rounded.minute
    if minutes < 30:
        return rounded.replace(minute=0)
    return rounded.replace(minute=30)

def get_next_slot_start(current):
    """Get the start of the next 30-minute slot."""
    return current + timedelta(minutes=30)

def get_week_start(date):
    """Get the Sunday of the week for the given date."""
    delta = (date.weekday() + 1) % 7
    return date - timedelta(days=delta)

def get_or_create_sheet(wb, week_start):
    """Get or create an Excel sheet for the week starting on week_start."""
    sheet_name = f"Week_{week_start.strftime('%Y-%m-%d')}"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # Headers
        ws.cell(1, 1, "Weekly Schedule")
        ws.cell(2, 1, "Time")
        ws.cell(2, 3, "Duration")
        ws.cell(2, 4, "Sunday")
        ws.cell(2, 5, "Monday")
        ws.cell(2, 6, "Tuesday")
        ws.cell(2, 7, "Wednesday")
        ws.cell(2, 8, "Thursday")
        ws.cell(2, 9, "Friday")
        ws.cell(2, 10, "Saturday")
        ws.cell(3, 1, "From")
        ws.cell(3, 2, "End")
        # Serial dates
        for d in range(7):
            day = week_start + timedelta(days=d)
            serial = (day - datetime(1899, 12, 30).date()).days
            ws.cell(3, 4 + d, serial)
        # Time slots
        for i in range(48):
            r = 4 + i
            from_t = i / 48.0
            end_t = (i + 1) / 48.0
            dur = 0.5
            ws.cell(r, 1, from_t)
            ws.cell(r, 2, end_t)
            ws.cell(r, 3, dur)
        return ws
    return wb[sheet_name]

def log_to_excel(log_date, slot_idx, activity):
    """Log an activity to the Excel file."""
    for attempt in range(2):
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
            week_start = get_week_start(log_date)
            ws = get_or_create_sheet(wb, week_start)
            col = 4 + (log_date - week_start).days
            row = 4 + slot_idx
            ws.cell(row, col, activity)
            wb.save(FILE_PATH)
            logging.info(f"Logged to Excel: {activity} at {log_date} slot {slot_idx}")
            return
        except Exception as e:
            logging.error(f"Excel save attempt {attempt + 1} failed: {str(e)}")
            if attempt == 1:
                try:
                    bot.send_message(CHAT_ID, f"Error saving Excel after retry: {str(e)}")
                except Exception as send_e:
                    logging.error(f"Failed to send Excel error message: {str(send_e)}")
            time.sleep(1)

def parse_number(text):
    """Parse numeric expressions like '1.5', '1 and a half', '90', '1 1/2'."""
    text = text.lower().strip()
    word_map = {
        'zero': 0,
        'one': 1,
        'two': 2,
        'three': 3,
        'four': 4,
        'five': 5,
        'six': 6,
        'seven': 7,
        'eight': 8,
        'nine': 9,
        'ten': 10,
        'half': 0.5,
        'quarter': 0.25,
    }
    if text in word_map:
        return word_map[text]
    try:
        return float(text)
    except ValueError:
        pass
    m = re.match(r'(\w+)\s+and\s+a?\s*(half|quarter)', text)
    if m:
        base = parse_number(m.group(1))
        frac = word_map[m.group(2)]
        if base is not None:
            return base + frac
    m = re.match(r'(\d+)\s+(\d+)/(\d+)', text)
    if m:
        return int(m.group(1)) + int(m.group(2)) / int(m.group(3))
    m = re.match(r'(\d+)/(\d+)', text)
    if m:
        return int(m.group(1)) / int(m.group(2))
    return None

def parse_response(text):
    """Parse response for activity and duration, defaulting to 1 slot."""
    match = re.match(r'(.+?)\s*for\s*(.+?)\s*(hour|minute)s?$', text, re.IGNORECASE)
    if match:
        act = match.group(1).strip()
        num_text = match.group(2).strip()
        unit = match.group(3).lower()
        num = parse_number(num_text)
        if num is not None:
            mins = num * 60 if 'hour' in unit else num
            num_slots = max(1, math.ceil(mins / 30))
            return act, num_slots
    return text.strip(), 1

def parse_modify_response(text):
    """Parse modification response for past slots in format 'For [start]-[end]: [act]'."""
    match = re.match(r'(?i)for\s+(?P<start>\d{1,2}:\d{2}\s*(?:AM|PM)?)\s*[-–]\s*(?P<end>\d{1,2}:\d{2}\s*(?:AM|PM)?)\s*:\s*(?P<act>.*)', text)
    if match:
        start_str = match.group('start').strip()
        end_str = match.group('end').strip()
        act = match.group('act').strip()
        # Parse times with optional space before AM/PM
        def parse_time(t_str):
            for fmt in ['%I:%M%p', '%I:%M %p']:
                try:
                    return datetime.strptime(t_str, fmt).time()
                except ValueError:
                    pass
            return None
        start_time = parse_time(start_str)
        end_time = parse_time(end_str)
        if not start_time or not end_time:
            return None, None, None
        # Calculate duration
        start_d = datetime(2000, 1, 1, start_time.hour, start_time.minute)
        end_d = datetime(2000, 1, 1, end_time.hour, end_time.minute)
        if end_d < start_d:
            end_d += timedelta(days=1)
        duration = (end_d - start_d).total_seconds()
        if duration <= 0 or duration % 1800 != 0:
            return None, None, None
        return start_time, end_time, act
    return None, None, None

def timeout():
    """Handle slot timeout if no response was received."""
    global has_response, current_slot_start, last_slot_start
    if current_slot_start:
        end_t = current_slot_start + timedelta(minutes=30)
        if not has_response:
            slot_idx = current_slot_start.hour * 2 + (1 if current_slot_start.minute == 30 else 0)
            log_to_excel(current_slot_start.date(), slot_idx, "No response")
            try:
                bot.send_message(CHAT_ID, f"Got it: No response for {current_slot_start:%I:%M %p}-{end_t:%I:%M %p}")
                logging.info(f"Logged no response for {current_slot_start:%I:%M %p}-{end_t:%I:%M %p}")
            except Exception as e:
                logging.error(f"Failed to send no response message: {str(e)}")
    last_slot_start = current_slot_start
    current_slot_start = None
    has_response = False

def ask(slot_start):
    """Send a query for the given slot and schedule timers."""
    global current_slot_start, last_slot_start, has_response, timeout_timer
    current_slot_start = slot_start
    last_slot_start = slot_start
    has_response = False
    end_t = slot_start + timedelta(minutes=30)
    try:
        bot.send_message(CHAT_ID, f"Hey Mo! What are you doing for the upcoming half hour? ({slot_start:%I:%M %p} - {end_t:%I:%M %p}) 📅")
        logging.info(f"Asked for activity at {slot_start:%I:%M %p}")
    except Exception as e:
        logging.error(f"Failed to send ask message: {str(e)}")
    # Schedule timeout relative to slot start
    now = datetime.now(TZ)
    timeout_time = (slot_start + timedelta(minutes=30) - now).total_seconds()
    if timeout_time > 0:
        timeout_timer = threading.Timer(timeout_time, timeout)
        timeout_timer.start()

@bot.message_handler(func=lambda message: True)
def handle_response(message):
    """Handle user responses, allowing overrides within the slot or modifications for past slots."""
    if message.chat.id != CHAT_ID:
        return
    text = message.text.strip()
    now = datetime.now(TZ)
    # Check if it's a modification for past slot
    modify_start_time, modify_end_time, modify_act = parse_modify_response(text)
    if modify_start_time:
        # Calculate start_dt
        start_dt = now.replace(hour=modify_start_time.hour, minute=modify_start_time.minute, second=0, microsecond=0)
        if start_dt > now:
            start_dt -= timedelta(days=1)
        # Calculate end_dt
        end_dt = now.replace(hour=modify_end_time.hour, minute=modify_end_time.minute, second=0, microsecond=0)
        if end_dt > now:
            end_dt -= timedelta(days=1)
        if end_dt <= start_dt:
            end_dt += timedelta(days=1)
        duration = (end_dt - start_dt).total_seconds()
        num_slots = int(duration / 1800)
        for i in range(num_slots):
            slot_dt = start_dt + timedelta(minutes=30 * i)
            slot_idx = slot_dt.hour * 2 + (1 if slot_dt.minute == 30 else 0)
            log_to_excel(slot_dt.date(), slot_idx, modify_act)
        try:
            bot.reply_to(message, f"Got it: {modify_act} for {start_dt:%I:%M %p}-{end_dt:%I:%M %p}")
            logging.info(f"Modified activity: {modify_act} for {start_dt:%I:%M %p}-{end_dt:%I:%M %p}")
        except Exception as e:
            logging.error(f"Failed to send modification confirmation: {str(e)}")
        return
    # Otherwise, handle as normal response for current slot, allowing overrides during multi-slots
    global has_response, multi_end, timeout_timer, current_slot_start, last_slot_start
    GRACE_PERIOD = timedelta(minutes=5)
    if current_slot_start is None:
        if last_slot_start and now < last_slot_start + GRACE_PERIOD:
            current_slot_start = last_slot_start
            has_response = False
        elif multi_end is None or now >= multi_end:
            try:
                bot.reply_to(message, "No active query to respond to.")
                logging.info("Received response with no active query")
            except Exception as e:
                logging.error(f"Failed to send no active query message: {str(e)}")
            return
        else:
            current_slot_start = get_current_slot_start(now)
            has_response = False
    # Allow responses until timeout fires
    slot_end = current_slot_start + timedelta(minutes=30)
    if now >= slot_end and (multi_end is None or now < multi_end):
        try:
            bot.reply_to(message, "Sorry, too late! Already logged as no response.")
            logging.info("Response received too late")
        except Exception as e:
            logging.error(f"Failed to send too late message: {str(e)}")
        return
    # Cancel existing timeout if any
    if timeout_timer:
        timeout_timer.cancel()
    activity, num_slots = parse_response(text)
    # Clear previous multi-slot entries beyond current slot
    if multi_end and multi_end > current_slot_start + timedelta(minutes=30):
        for i in range(1, 48):  # Max 24 hours
            slot_dt = current_slot_start + timedelta(minutes=30 * i)
            if slot_dt >= multi_end:
                break
            slot_idx = slot_dt.hour * 2 + (1 if slot_dt.minute == 30 else 0)
            log_to_excel(slot_dt.date(), slot_idx, "")
    # Log new activity
    for i in range(num_slots):
        slot_dt = current_slot_start + timedelta(minutes=30 * i)
        slot_idx = slot_dt.hour * 2 + (1 if slot_dt.minute == 30 else 0)
        log_to_excel(slot_dt.date(), slot_idx, activity)
    has_response = True
    end_t = current_slot_start + timedelta(minutes=30 * num_slots)
    try:
        if num_slots > 1:
            bot.reply_to(message, f"Got it: {activity} from {current_slot_start:%I:%M %p} to {end_t:%I:%M %p} 🚀")
            logging.info(f"Logged multi-slot activity: {activity} from {current_slot_start:%I:%M %p} to {end_t:%I:%M %p}")
            multi_end = end_t
        else:
            bot.reply_to(message, f"Got it: {activity} for {current_slot_start:%I:%M %p}-{end_t:%I:%M %p}")
            logging.info(f"Logged activity: {activity} for {current_slot_start:%I:%M %p}-{end_t:%I:%M %p}")
            multi_end = None  # Clear multi_end for single-slot responses
    except Exception as e:
        logging.error(f"Failed to send confirmation message: {str(e)}")
    # No reminder, but reschedule timeout if single-slot and time left
    if num_slots == 1:
        timeout_time = (current_slot_start + timedelta(minutes=30) - datetime.now(TZ)).total_seconds()
        if timeout_time > 0:
            timeout_timer = threading.Timer(timeout_time, timeout)
            timeout_timer.start()

def polling_loop():
    """Main loop to trigger queries for each slot."""
    global current_slot_start
    while True:
        now = datetime.now(TZ)
        if multi_end and now < multi_end:
            sleep_sec = (multi_end - now).total_seconds()
            time.sleep(max(1, sleep_sec))
            continue
        slot_start = get_current_slot_start(now)
        if current_slot_start is None or slot_start > current_slot_start:
            ask(slot_start)
        next_slot = get_next_slot_start(slot_start)
        sleep_sec = (next_slot - now).total_seconds()
        if sleep_sec > 0:
            time.sleep(sleep_sec)

def run_bot():
    """Run the bot with retry logic for network errors."""
    while True:
        try:
            logging.info("Starting bot polling")
            bot.polling(none_stop=True, timeout=60)
        except Exception as e:
            logging.error(f"Polling crashed: {str(e)}")
            time.sleep(10)
            bot.stop_polling()
            logging.info("Retrying polling")

if __name__ == '__main__':
    threading.Thread(target=polling_loop, daemon=True).start()
    run_bot()
